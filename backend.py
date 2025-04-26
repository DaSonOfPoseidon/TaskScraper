import os
import re
import time
import pickle
from datetime import datetime
from tkinter import Tk, simpledialog, messagebox
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support import expected_conditions as EC
from dotenv import load_dotenv, set_key
from tqdm import tqdm
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from collections import defaultdict
from threading import Lock
from concurrent.futures import ThreadPoolExecutor


cookie_lock = Lock()
TASK_URL = "http://inside.sockettelecom.com/menu.php?tabid=45&tasktype=12&nDepartmentID=1&width=1440&height=731"
MAX_THREADS = 5  # global, easy to change later
PAGE_TIMEOUT = 30  # seconds to allow a page to load

# === Login & Session ===
def prompt_for_credentials():
    login_window = Tk()
    login_window.withdraw()
    USERNAME = simpledialog.askstring("Login", "Enter your USERNAME:", parent=login_window)
    PASSWORD = simpledialog.askstring("Login", "Enter your PASSWORD:", parent=login_window, show="*")
    login_window.destroy()
    return USERNAME, PASSWORD

def save_env_credentials(USERNAME, PASSWORD):
    dotenv_path = ".env"
    if not os.path.exists(dotenv_path):
        with open(dotenv_path, "w") as f:
            f.write("")
    set_key(dotenv_path, "UNITY_USER", USERNAME)
    set_key(dotenv_path, "PASSWORD", PASSWORD)

def check_env_or_prompt_login(log=print):
    load_dotenv()
    username = os.getenv("UNITY_USER")
    password = os.getenv("PASSWORD")
    if username and password:
        log("✅ Loaded stored credentials.")
        return username, password
    while True:
        username, password = prompt_for_credentials()
        if not username or not password:
            messagebox.showerror("Login Cancelled", "Login is required to continue.")
            return None, None
        save_env_credentials(username, password)
        log("✅ Credentials captured and saved to .env.")
        return username, password

def save_cookies(driver, filename="cookies.pkl"):
    with cookie_lock:
        with open(filename, "wb") as f:
            pickle.dump(driver.get_cookies(), f)

def load_cookies(driver, filename="cookies.pkl"):
    if not os.path.exists(filename): return False
    try:
        with cookie_lock:
            with open(filename, "rb") as f:
                cookies = pickle.load(f)
        driver.get("http://inside.sockettelecom.com/")
        for cookie in cookies:
            driver.add_cookie(cookie)
        driver.refresh()
        clear_first_time_overlays(driver)
        return True
    except Exception:
        if os.path.exists(filename): os.remove(filename)
        return False

def handle_login(driver, log=print):
    driver.get("http://inside.sockettelecom.com/")
    bypass_ssl_warning(driver)
    if load_cookies(driver):
        if not login_failed(driver):
            log("✅ Session restored via cookies.")
            clear_first_time_overlays(driver)
            return
        else:
            print("⚠️ Cookie session invalid — retrying with credentials...")
    while "login.php" in driver.current_url or "Username" in driver.page_source:
        username, password = check_env_or_prompt_login(log)
        if not username or not password:
            log("❌ Login cancelled.")
            return
        perform_login(driver, username, password)
        WebDriverWait(driver, 10).until(lambda d: "menu.php" in d.current_url)
        if not login_failed(driver):
            save_cookies(driver)
            log("✅ Logged in with username/password.")
            return
        else:
            log("❌ Login failed. Re-prompting...")

def perform_login(driver, USERNAME, PASSWORD):
    driver.get("http://inside.sockettelecom.com/system/login.php")
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.NAME, "username")))
    driver.find_element(By.NAME, "username").send_keys(USERNAME)
    driver.find_element(By.NAME, "password").send_keys(PASSWORD)
    driver.find_element(By.ID, "login").click()
    clear_first_time_overlays(driver)

def login_failed(driver):
    try:
        return (
            "login.php" in driver.current_url
            or "Username" in driver.page_source
            or "Invalid username or password" in driver.page_source
        )
    except Exception:
        return True

def clear_first_time_overlays(driver):
    try:
        WebDriverWait(driver, 0.5).until(EC.alert_is_present())
        driver.switch_to.alert.dismiss()
    except:
        pass
    buttons = [
        "//form[@id='valueForm']//input[@type='button']",
        "//form[@id='f']//input[@type='button']"
    ]
    for xpath in buttons:
        try:
            WebDriverWait(driver, 0.5).until(EC.element_to_be_clickable((By.XPATH, xpath))).click()
        except:
            pass
    for _ in range(3):
        try:
            WebDriverWait(driver, 0.5).until(EC.frame_to_be_available_and_switch_to_it((By.ID, "MainView")))
            return
        except:
            time.sleep(0.25)

def extract_sales_order_id(task):
    match = re.search(r"OrderID=(\d+)", task["view_url"])
    return match.group(1) if match else None

def extract_job_name(description):
    if "):" in description:
        return description.split("):", 1)[-1].strip()
    return description.strip()

def normalize_job_name(name):
    name = name.lower()
    if "install" in name and "on-site" in name:
        return "on-site install"
    return name

def extract_name_and_cid(company_text):
    match = re.search(r"(.*?)\s*-\s*(\d{4}-\d{4}-\d{4})", company_text)
    if match:
        return f"{match.group(1).strip()} - {match.group(2).strip()}"
    return company_text.strip()

def parse_task_row(row):
    try:
        cells = row.find_elements(By.TAG_NAME, "td")
        if not cells or len(cells) < 6:
            return None
        view_link = row.find_element(By.CSS_SELECTOR, "a.button").get_attribute("href")
        description = cells[1].text.strip()
        date_assigned = cells[2].text.strip()
        due_date = cells[3].text.strip()
        assigned = cells[4].text.strip()
        company = cells[5].text.strip()
        print(f"Customer {company} task parsed")
        return {
            "view_url": view_link,
            "description": description,
            "date_assigned": date_assigned,
            "due_date": due_date,
            "assigned": assigned,
            "company": company
        }
    except Exception as e:
        print(f"[!] Skipping row due to error: {e}")
        return None

def scrape_department_tasks(driver):
    driver.execute_script(f"window.location.href = '{TASK_URL}'")
    time.sleep(2)
    bypass_ssl_warning(driver)
    WebDriverWait(driver, 60).until(
        EC.frame_to_be_available_and_switch_to_it((By.ID, "MainView"))
    )
    print("✅ Switched to MainView iframe. Waiting for task rows...")
    rows = []
    for _ in range(60):
        rows = driver.find_elements(By.XPATH, '//tr[contains(@class, "taskElement")]')
        if rows:
            print(f"✅ Found {len(rows)} task rows.")
            break
        time.sleep(1)
    else:
        raise TimeoutException("❌ Timed out waiting for taskElement rows in iframe.")

    tasks = []
    with ThreadPoolExecutor(max_workers=20) as executor:
        futures = [executor.submit(parse_task_row, row) for row in rows]
        for future in futures:
            result = future.result()
            if result:
                tasks.append(result)
    return tasks

def prompt_for_job_types(task_list):
    job_names = sorted({extract_job_name(task["description"]) for task in task_list})
    print("\nAvailable job types:")
    for i, job in enumerate(job_names, 1):
        print(f"{i}. {job}")
    selected_indices = input("\nEnter the numbers of the job types you'd like to include (comma-separated): ")
    selected = set()
    try:
        for i in selected_indices.split(","):
            selected.add(job_names[int(i.strip()) - 1])
    except:
        print("❌ Invalid input, defaulting to all.")
        return job_names
    return selected

def filter_and_sort_tasks(tasks, selected_job_types):
    tasks_by_so = defaultdict(list)
    for task in tasks:
        so_id = extract_sales_order_id(task)
        if so_id:
            tasks_by_so[so_id].append(task)

    final_tasks = batch_check_sales_orders(tasks_by_so, selected_job_types)

    def parse_due_date(task):
        try:
            return datetime.strptime(task["due_date"], "%Y-%m-%d %H:%M:%S")
        except:
            return datetime.max
    return sorted(final_tasks, key=parse_due_date)

def wait_for_so_page(driver, timeout=30): # checks for laoding screen
    try:
        WebDriverWait(driver, timeout, poll_frequency=0.2).until(
            lambda d: d.execute_script("""
                try {
                    var iframe = document.getElementById('MainView');
                    if (!iframe) return false;
                    var innerDoc = iframe.contentDocument || iframe.contentWindow.document;
                    return innerDoc && innerDoc.querySelector('h3.taskName') !== null;
                } catch (e) {
                    return false;
                }
            """)
        )
        print("✅ MainView iframe loaded and SO page is ready.")
        return True
    except TimeoutException:
        print("⚠️ Timeout waiting for SO page to fully load.")
        return False

def export_tasks_to_excel(tasks):
    today = datetime.today().date()
    output_dir = os.path.join("Outputs")
    os.makedirs(output_dir, exist_ok=True)
    filename = os.path.join(output_dir, f"{today.strftime('%m%d')}OpenTasks.xlsx")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Filtered Tasks"
    ws.append(["Name - CID", "Due Date", "Task Name", "SO LINK"])

    for task in tasks:
        # Parse due date
        try:
            due_date = datetime.strptime(task["due_date"], "%Y-%m-%d %H:%M:%S").date()
        except:
            due_date = today  # If parsing fails, assume today so it still shows

        if due_date <= today:
            name_cid = extract_name_and_cid(task["company"])
            task_name = extract_job_name(task["description"])
            so_url = task["view_url"]
            ws.append([name_cid, task["due_date"], task_name, so_url])
            cell = ws.cell(row=ws.max_row, column=4)
            cell.hyperlink = so_url
            cell.font = Font(color="0000FF", underline="single")

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 2, 15)

    wb.save(filename)
    print(f"\n✅ Task export complete. File saved as '{filename}'")

def bypass_ssl_warning(driver):
    try:
        time.sleep(1)
        if "privacy-error" in driver.page_source or "proceed-button" in driver.page_source:
            print("⚠️ SSL warning page detected. Attempting to click through...")
            WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.ID, "proceed-button"))
            ).click()
            print("✅ Clicked 'Continue to site' button.")
            time.sleep(2)
    except Exception as e:
        print(f"❌ SSL bypass failed: {e}")

def should_include_so(driver, view_url):
    try:
        driver.get(view_url)
        wait_for_so_page(driver)
        if check_incomplete_schedule_tasks(driver):
            print(f"⛔ Skipping {driver.current_url} — Incomplete scheduling tasks found")
            return False
        
        # Check onsite install date
        try:
            install_info = driver.find_element(By.ID, "onsiteInstallContainer").text
            match = re.search(r"(\d{4}-\d{2}-\d{2})", install_info)
            if match:
                install_date = datetime.strptime(match.group(1), "%Y-%m-%d").date()
                today = datetime.today().date()
                if install_date > today:
                    print(f"⛔ Skipping SO {so_id} — Install scheduled for {install_date}")
                    return False
        except:
            pass  # No install date block found — assume OK

        return True

    except Exception as e:
        print(f"⚠️ Error processing SO {view_url}: {e}")
        return False

def batch_check_sales_orders(tasks_by_so, selected_job_types):
    """Batch checks Sales Orders by splitting them across threads"""
    all_links = list(tasks_by_so.keys())
    total = len(all_links)

    print(f"🔎 Total Sales Orders to check: {total}")

    # Split SOs roughly evenly
    chunk_size = (total + MAX_THREADS - 1) // MAX_THREADS  # ceil division
    chunks = [all_links[i:i + chunk_size] for i in range(0, total, chunk_size)]

    results = []

    with ThreadPoolExecutor(max_workers=MAX_THREADS) as executor:
        futures = []
        for chunk in chunks:
            futures.append(executor.submit(check_multiple_sos_worker, chunk, selected_job_types))

        for future in tqdm(futures, desc="Checking Sales Orders", unit="SO"):
            try:
                chunk_results = future.result()
                results.extend(chunk_results)
            except Exception as e:
                print(f"⚠️ Error in thread: {e}")

    print(f"✅ {len(results)} Sales Orders passed validation.\n")
    return results

def check_incomplete_schedule_tasks(driver):
    try:
        # Switch into iframe if exists
        try:
            driver.switch_to.default_content()
            iframe = driver.find_element(By.NAME, "MainView")
            driver.switch_to.frame(iframe)
            print("🔎 Switched into 'MainView' iframe.")
        except Exception as e:
            print(f"⚠️ No 'MainView' iframe found, continuing in current context. ({e})")

        task_containers = driver.find_elements(By.CLASS_NAME, "taskContainer")
        incomplete_tasks = 0
        relevant_tasks = 0

        for container in task_containers:
            try:
                header = container.find_element(By.TAG_NAME, "h3")
                title = header.text.lower()

                if any(keyword in title for keyword in ["schedule", "reschedule", "contact customer"]):
                    relevant_tasks += 1
                    class_attr = header.get_attribute("class")
                    if "completed" not in class_attr:
                        print(f"❌ Incomplete task detected: '{title}' with class '{class_attr}'")
                        incomplete_tasks += 1
                    else:
                        print(f"✅ Completed task detected: '{title}' with class '{class_attr}'")
            except Exception as e:
                print(f"⚠️ Error checking task container: {e}")

        print(f"🔎 Checked {relevant_tasks} scheduling tasks: {relevant_tasks - incomplete_tasks} completed, {incomplete_tasks} incomplete.")
        return incomplete_tasks > 0
    except Exception as e:
        print(f"⚠️ Error during schedule/reschedule/contact task check: {e}")
        return True

def check_multiple_sos_worker(so_links, selected_job_types):
    """Each thread will open one driver and process multiple SOs"""
    driver = None
    results = []

    try:
        driver = create_driver()
        handle_login(driver)

        for link in so_links:
            result = check_single_so(driver, link, selected_job_types)
            if result:
                results.append(result)

    except Exception as e:
        print(f"⚠️ Worker error: {e}")

    finally:
        if driver:
            driver.quit()

    return results

def create_driver():
    """Creates a new Chrome WebDriver instance"""
    options = webdriver.ChromeOptions()
    options = webdriver.ChromeOptions()
    options.add_argument("--headless=new")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--allow-insecure-localhost")
    options.add_argument("--ignore-certificate-errors")
    options.add_argument("--allow-running-insecure-content")
    options.add_argument("--ignore-certificate-errors-spki-list")
    options.add_argument("--ignore-urlfetcher-cert-requests")
    options.add_argument("--ignore-ssl-errors=yes")
    options.add_argument("--test-type")
    options.page_load_strategy = 'eager'
    service = Service('chromedriver.exe')  # or your correct path
    driver = webdriver.Chrome(service=service, options=options)
    return driver

if __name__ == "__main__":
    options = webdriver.ChromeOptions()
    options.add_argument("--headless=new")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--allow-insecure-localhost")
    options.add_argument("--ignore-certificate-errors")
    options.add_argument("--allow-running-insecure-content")
    options.add_argument("--ignore-certificate-errors-spki-list")
    options.add_argument("--ignore-urlfetcher-cert-requests")
    options.add_argument("--ignore-ssl-errors=yes")
    options.add_argument("--test-type")
    options.page_load_strategy = 'eager'
    driver = webdriver.Chrome(options=options)
    handle_login(driver)
    all_tasks = scrape_department_tasks(driver)
    selected_types_raw = prompt_for_job_types(all_tasks)
    selected_types = set(normalize_job_name(name) for name in selected_types_raw)
    final_tasks = filter_and_sort_tasks(all_tasks, selected_types)
    export_tasks_to_excel(final_tasks)
